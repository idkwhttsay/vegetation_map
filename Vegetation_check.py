from PIL import Image
import openpyxl

image_path = '/Users/User/Downloads/photo_2023-09-03 05.25.41.jpeg'
image = Image.open(image_path)

width, height = image.size
green_pixel_count = 0

workbook = openpyxl.load_workbook("/Users/User/Desktop/green.xlsx")
sheet = workbook.active

sheet.cell(row=1, column=1, value="image")
sheet.cell(row=1, column=2, value="percentage")
sheet.cell(row=1, column=3, value="money")



for y in range(height):
    for x in range(width):
        pixel_color = image.getpixel((x, y))

        if pixel_color[1] > 10 and pixel_color[0] < 80 and pixel_color[2] < 80:
            green_pixel_count += 1

total_pixels = width*height
percent_green = (green_pixel_count/total_pixels) * 100

percentage = 40 - percent_green
money = percentage * 10 * 2000

if (percentage < 0):
    money = 0

input_data = ["павлодарская область"]
output_data = [percent_green]
data_data = [money]

for i, (input_value, output_value, data_value) in enumerate(zip(input_data, output_data,data_data), start=8):
    sheet.cell(row=i, column=1, value=input_value)
    sheet.cell(row=i, column=2, value=output_value)
    sheet.cell(row=i, column=3, value=data_value)

workbook.save("/Users/User/Desktop/green.xlsx")

print(f"Green(and its shades) pixels: {green_pixel_count} ({percent_green:.2f}%)")
if(percent_green < 40):
    print("Low vegetation zone")
else:
    print("High vegetation zone")
